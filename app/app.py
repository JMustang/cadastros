import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook
from tkinter import StringVar

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.sistema()

    def layout_config(self):
        self.title("Cadastros")
        self.geometry("700x500")
        self.resizable(False, False)  # Impede a redimensionamento horizontal e vertical

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(
            self, text="Tema", bg_color="transparent", text_color=["#000", "#fff"]
        ).place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(
            self, values=["Light", "Dark", "System"], command=self.change_apm
        ).place(x=50, y=460)

    def sistema(self):
        frame = ctk.CTkFrame(
            self,
            width=700,
            height=50,
            corner_radius=0,
            bg_color="darkmagenta",
            fg_color="darkmagenta",
        )
        frame.place(x=0, y=0)
        title = ctk.CTkLabel(
            frame,
            text="Cadastros",
            font=("Arial bold", 24),
            text_color="#aaa",
        ).place(x=275, y=10)
        span = ctk.CTkLabel(
            self,
            text="Por favor, preencha o formulário!",
            font=("Arial bold", 16),
            text_color=["#000", "#fff"],
        ).place(x=50, y=70)

        tabela = pathlib.Path("Cadastros.xlsx")
        if tabela.exists():
            pass
        else:
            tabela = Workbook()
            tabela.save("Cadastros.xlsx")
            tabela = openpyxl.load_workbook("Cadastros.xlsx")
            data = tabela.active
            data["A1"] = "Nome completo"
            data["B1"] = "Contato"
            data["C1"] = "Idade"
            data["D1"] = "Genero"
            data["E1"] = "Endereco"
            data["F1"] = "Obcervacoes"
            tabela.save(r"Cadastros.xlsx")
            messagebox.showinfo("Sistema", "Arquivo criado com sucesso!")

        # Funcao dos botoes
        def salva():
            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            address = address_value.get()
            gender = opt_gender.get()
            obs = txt_opt.get(0.0, END)

            if (
                name == ""
                or contact == ""
                or age == ""
                or address == ""
                or gender == ""
                or obs == ""
            ):
                messagebox.showerror("Sistema", "ERROR!\nPreencha todos os campos!")
                return

            tabela = openpyxl.load_workbook("Cadastros.xlsx")
            data = tabela.active
            data.cell(row=data.max_row + 1, column=1, value=name)
            data.cell(row=data.max_row, column=2, value=contact)
            data.cell(row=data.max_row, column=3, value=age)
            data.cell(row=data.max_row, column=4, value=address)
            data.cell(row=data.max_row, column=5, value=gender)
            data.cell(row=data.max_row, column=6, value=obs)

            tabela.save(r"Cadastros.xlsx")
            messagebox.showinfo("Sistema", "Dados salvos com sucesso!")

        def limpar():
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            address_value.set("")
            txt_opt.delete(0.0, END)

        # Texts variables
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()

        # Entrys
        en_name = ctk.CTkEntry(
            self,
            width=350,
            textvariable=name_value,
            font=("Century Gohtic bold", 16),
            fg_color="transparent",
            placeholder_text="Nome",
        )
        en_contact = ctk.CTkEntry(
            self,
            width=200,
            textvariable=contact_value,
            font=("Century Gohtic bold", 16),
            fg_color="transparent",
            placeholder_text="Contato",
        )
        en_age = ctk.CTkEntry(
            self,
            width=150,
            textvariable=age_value,
            font=("Century Gohtic bold", 16),
            fg_color="transparent",
            placeholder_text="Idade",
        )
        en_address = ctk.CTkEntry(
            self,
            width=200,
            textvariable=address_value,
            font=("Century Gohtic bold", 16),
            fg_color="transparent",
            placeholder_text="Endereço",
        )

        # Combobox
        opt_gender = ctk.CTkComboBox(
            self,
            values=["Masculino", "Feminino"],
            font=("Century Gohtic bold", 14),
            width=150,
        )
        opt_gender.set("Masculino")

        # TextBox
        txt_opt = ctk.CTkTextbox(
            self,
            width=500,
            height=150,
            font=("Arial", 18),
            border_color="#aaa",
            border_width=2,
            fg_color="transparent",
        )

        # Labels
        lb_name = ctk.CTkLabel(
            self,
            text="Nome:",
            font=("Arial bold", 16),
            text_color=["#000", "#fff"],
        )
        lb_contact = ctk.CTkLabel(
            self,
            text="Contato:",
            font=("Arial bold", 16),
            text_color=["#000", "#fff"],
        )
        lb_age = ctk.CTkLabel(
            self,
            text="Idade:",
            font=("Arial bold", 16),
            text_color=["#000", "#fff"],
        )
        lb_gender = ctk.CTkLabel(
            self,
            text="Genero:",
            font=("Arial bold", 16),
            text_color=["#000", "#fff"],
        )
        lb_address = ctk.CTkLabel(
            self,
            text="endereço:",
            font=("Arial bold", 16),
            text_color=["#000", "#fff"],
        )
        lb_obs = ctk.CTkLabel(
            self,
            text="observação:",
            font=("Arial bold", 16),
            text_color=["#000", "#fff"],
        )

        # Botões
        btn_save = ctk.CTkButton(
            self,
            text="Salva dados".upper(),
            command=salva,
            fg_color="#151",
            hover_color="#131",
        ).place(x=300, y=420)
        btn_clear = ctk.CTkButton(
            self,
            text="Limpar campos".upper(),
            command=limpar,
            fg_color="#555",
            hover_color="#333",
        ).place(x=500, y=420)

        # Posicionando
        lb_name.place(x=50, y=120)
        en_name.place(x=50, y=150)

        lb_contact.place(x=450, y=120)
        en_contact.place(x=450, y=150)

        lb_age.place(x=300, y=190)
        en_age.place(x=300, y=220)

        lb_gender.place(x=500, y=190)
        opt_gender.place(x=500, y=220)

        lb_address.place(x=50, y=190)
        en_address.place(x=50, y=220)

        lb_obs.place(x=50, y=260)
        txt_opt.place(x=150, y=260)

    def change_apm(self, theme):
        ctk.set_appearance_mode(theme)
        self.appearence()


if __name__ == "__main__":
    app = App()
    app.mainloop()
